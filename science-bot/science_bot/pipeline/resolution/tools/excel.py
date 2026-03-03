"""Excel-specific inspection tools for resolution."""

import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl

from science_bot.pipeline.resolution.tools.reader import parse_filename


def list_excel_sheets(capsule_path: Path, filename: str) -> list[str]:
    """Return sheet names for an Excel file.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Direct filename or zip-contained Excel path.

    Returns:
        list[str]: Sheet names in workbook order.
    """
    ref = parse_filename(capsule_path, filename)
    if ref.zip_path is not None and ref.inner_path is not None:
        with zipfile.ZipFile(ref.zip_path) as archive:
            data = archive.read(ref.inner_path)
        workbook = openpyxl.load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
        )
    else:
        workbook = openpyxl.load_workbook(
            ref.file_path,
            read_only=True,
            data_only=True,
        )
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names
